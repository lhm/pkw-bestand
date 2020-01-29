from utils import root, filesdir, datadir
import openpyxl
import pandas as pd

headers = [
    'Zulassungsbezirk',
    'Benzin',
    'Diesel',
    'Gas',
    'Hybrid',
    'Elektro',
    'Sonstige Kraftstoffart',
    'Euro 1',
    'Euro 2',
    'Euro 3',
    'Euro 4',
    'Euro 5',
    'Euro 6',
    'Sonstige Emissionsgruppe'
 ]

args = {
    'sheet_name': 'Pkw',
    'header': None,
    'names': headers,
    'usecols': 'D,F:I,K:L,M:R,U',
    'engine': 'openpyxl',
    'skiprows': list(range(0, 9)),
    'skipfooter': 5,
    'thousands': '.',
    'na_values': '-'
}

infile = filesdir / 'fz1_2019_xlsx.xlsx'
outfile = datadir / 'kfzbestand-20190101.csv'

def clean(df):
    # remove subtotal rows
    subtotals = pd.isna(df.Zulassungsbezirk)
    df = df[~subtotals]

    # split Zulassungsbezirksname & Kreisschlüssel into separate columns
    zb = df.Zulassungsbezirk.str.strip().str.split('  ', expand=True)
    df = df.assign(KS=zb[0], Zulassungsbezirk=zb[1])
    df = df[['KS'] + headers]
    return df

def validate(infile, outfile):
    df = pd.read_csv(outfile, dtype={'KS': 'string'})
    ws = openpyxl.load_workbook(infile)['Pkw']

    # compare totals in germany, ignoring geographically unaccounted ('Sonstige')
    total_by_krafstoffe_expected = ws['E456'].value - ws['E455'].value
    total_by_kraftstoffe_actual = df.iloc[:,2:8].sum().sum()
    assert(total_by_krafstoffe_expected == total_by_kraftstoffe_actual)

    total_by_emissionsgruppe_expected = ws['V456'].value - ws['V455'].value
    total_by_emissionsgruppe_actual = df.iloc[:,8:16].sum().sum()
    assert(total_by_emissionsgruppe_expected == total_by_emissionsgruppe_actual)

    # compare subtotal of sachsen using filter on KS
    elektro_sachsen_expected = ws['K399'].value
    elektro_sachsen_actual = df[df.KS.str.startswith('14')].iloc[:,6].sum()
    assert(elektro_sachsen_expected == elektro_sachsen_actual)

def main():
    df = pd.read_excel(infile, **args)
    df = clean(df)
    df.to_csv(outfile, index=False)
    validate(infile, outfile)

main()
